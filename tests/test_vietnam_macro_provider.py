from __future__ import annotations

import sqlite3
from dataclasses import replace
from datetime import date, datetime, timedelta, timezone
from io import BytesIO

import openpyxl
import pytest
import requests

from tradingagents.dataflows.vietnam_macro import (
    MAX_HTML_BYTES,
    NSO_DATASETS,
    MacroObservation,
    MacroStatus,
    VietnamMacroClient,
    VietnamMacroConfig,
    VietnamMacroResult,
    VietnamMacroService,
    _apply_staleness,
    _derived_growth,
    _discover_latest_release,
    _FetchResult,
    _parse_excel,
    _parse_nso_release,
    _parse_sbv_html,
    _parse_sdmx,
    render_vietnam_macro_result,
)
from tradingagents.dataflows.vietnam_macro_archive import (
    ArchiveConfigurationError,
    VietnamMacroArchive,
)

UTC = timezone.utc


def _config(tmp_path, **overrides):
    base = VietnamMacroConfig(
        enabled=True,
        providers=("nso_sdmx", "nso_release", "sbv_html"),
        lookback_months=24,
        strict_point_in_time=True,
        timeout_seconds=15,
        archive_path=tmp_path / "macro" / "vn_macro.sqlite3",
        max_retries=3,
    )
    return replace(base, **overrides)


def _observation(
    value: str,
    first_seen: datetime,
    *,
    published: datetime | None = None,
    indicator: str = "vn_core_cpi_yoy",
) -> MacroObservation:
    published = published or first_seen - timedelta(days=1)
    return MacroObservation(
        indicator_id=indicator,
        value=value,
        unit="percent",
        unit_multiplier=1,
        frequency="M",
        period_start=datetime(2025, 12, 1, tzinfo=UTC),
        period_end=datetime(2025, 12, 31, 23, 59, 59, tzinfo=UTC),
        published_at=published,
        first_seen_at=first_seen,
        retrieved_at=first_seen,
        source_provider="nso_sdmx",
        source_series="PCPICO_PC_PP_PT",
        source_url=NSO_DATASETS[0].xml_url,
        provisional=None,
        point_in_time_quality="proxy",
        derived_from=[],
    )


def _fetch(fetch_id: str, observation: MacroObservation) -> _FetchResult:
    return _FetchResult(
        fetch_id=fetch_id,
        provider="nso_sdmx",
        source_id="nso_cpi",
        source_url=NSO_DATASETS[0].xml_url,
        status=MacroStatus.AVAILABLE,
        observations=[observation],
        started_at=observation.retrieved_at - timedelta(seconds=1),
        completed_at=observation.retrieved_at,
        http_status=200,
        request_succeeded=True,
        response_hash=fetch_id,
    )


def _sdmx(dataset, series_blocks: str, prepared="2026-08-12T10:40:10Z") -> bytes:
    return (
        f"""<?xml version='1.0' encoding='UTF-8'?>
        <message:StructureSpecificData
          xmlns:message='http://www.sdmx.org/resources/sdmxml/schemas/v2_1/message'>
          <message:Header><message:Prepared>{prepared}</message:Prepared></message:Header>
          <DataSet>{series_blocks}</DataSet>
        </message:StructureSpecificData>"""
    ).encode()


def _series(domain, indicator, frequency, multiplier, observations, base=None):
    base_attr = f' BASE_PER="{base}"' if base is not None else ""
    body = "".join(
        f'<Obs TIME_PERIOD="{period}" OBS_VALUE="{value}"/>' for period, value in observations
    )
    return (
        f'<Series DATA_DOMAIN="{domain}" REF_AREA="VN" INDICATOR="{indicator}" '
        f'COUNTERPART_AREA="_Z" FREQ="{frequency}" UNIT_MULT="{multiplier}"'
        f"{base_attr}>{body}</Series>"
    )


@pytest.mark.unit
def test_sdmx_maps_exact_series_frequency_units_and_prepared_time():
    dataset = NSO_DATASETS[0]
    payload = _sdmx(
        dataset,
        _series("CPI", "PCPI_IX", "M", 0, [("2025-09", "120.1")], "2019")
        + _series("CPI", "PCPICO_PC_PP_PT", "M", 0, [("2025-09", "3.17")], "_Z"),
    )
    observed = datetime(2026, 8, 18, 5, tzinfo=UTC)

    observations, releases, warnings = _parse_sdmx(
        payload,
        dataset=dataset,
        observed_at=observed,
        source_url=dataset.xml_url,
    )

    assert releases == []
    assert warnings == []
    assert {(item.indicator_id, item.value) for item in observations} == {
        ("vn_cpi_index", "120.1"),
        ("vn_core_cpi_yoy", "3.17"),
    }
    assert all(
        item.published_at == datetime(2026, 8, 12, 10, 40, 10, tzinfo=UTC) for item in observations
    )
    assert all(item.first_seen_at == observed for item in observations)


@pytest.mark.unit
def test_sdmx_mapping_change_fails_closed():
    dataset = NSO_DATASETS[2]
    payload = _sdmx(
        dataset,
        _series("IND", "AIP_ISIC4_IX", "M", 3, [("2026-07", "214.6")], "2015"),
    )

    with pytest.raises(ValueError, match="unit multiplier changed"):
        _parse_sdmx(
            payload,
            dataset=dataset,
            observed_at=datetime.now(UTC),
            source_url=dataset.xml_url,
        )


@pytest.mark.unit
def test_sdmx_rejects_non_vietnam_reference_area():
    dataset = NSO_DATASETS[2]
    payload = _sdmx(
        dataset,
        _series("IND", "AIP_ISIC4_IX", "M", 0, [("2026-07", "214.6")], "2015").replace(
            'REF_AREA="VN"', 'REF_AREA="US"'
        ),
    )

    with pytest.raises(ValueError, match="reference area changed"):
        _parse_sdmx(
            payload,
            dataset=dataset,
            observed_at=datetime.now(UTC),
            source_url=dataset.xml_url,
        )


@pytest.mark.unit
def test_sdmx_conflicting_period_uses_final_occurrence_and_marks_partial():
    dataset = NSO_DATASETS[1]
    payload = _sdmx(
        dataset,
        _series(
            "NAG",
            "NGDP_R_PA_XDC",
            "Q",
            9,
            [
                ("2025-Q1", "1503311.3"),
                ("2025-Q1", "2401936.8"),
                ("2025-Q2", "2633501.7"),
            ],
            "2010",
        ),
    )

    observations, _releases, warnings = _parse_sdmx(
        payload,
        dataset=dataset,
        observed_at=datetime.now(UTC),
        source_url=dataset.xml_url,
    )

    assert len(observations) == 2
    q1 = next(item for item in observations if item.period_start.month == 1)
    assert q1.value == "2401936.8"
    assert q1.point_in_time_quality == "partial"
    assert any("conflicting values" in warning for warning in warnings)


def _workbook_bytes(sheets):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for title, metadata, header, rows in sheets:
        worksheet = workbook.create_sheet(title)
        for key, value in metadata:
            worksheet.append([key, value])
        for column, value in enumerate(header, start=1):
            worksheet.cell(row=11, column=column, value=value)
        for row_index, row in enumerate(rows, start=12):
            for column, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=column, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


@pytest.mark.unit
def test_excel_prefilters_frequency_and_uses_final_exact_sheet():
    dataset = NSO_DATASETS[1]
    metadata = [
        ("DATA_DOMAIN", "NAG"),
        ("REF_AREA", "VN"),
        ("COUNTERPART_AREA", "_Z"),
        ("UNIT_MULT", 9),
    ]
    header = [
        "Country code",
        "Descriptor",
        "Descriptor Vietnamese",
        "INDICATOR",
        "FREQ",
        "BASE_PER",
        "2025-Q1",
        "2025-Q2",
    ]
    payload = _workbook_bytes(
        [
            (
                "annual",
                metadata,
                [
                    "Country code",
                    "Descriptor",
                    "Descriptor Vietnamese",
                    "INDICATOR",
                    "FREQ",
                    "BASE_PER",
                    "2025",
                ],
                [
                    [
                        "NGDP_R_PA_XDC",
                        "Gross Domestic Product",
                        "Tổng sản phẩm theo giá so sánh 2010",
                        "NGDP_R_PA_XDC",
                        "A",
                        "2010",
                        999,
                    ]
                ],
            ),
            (
                "quarterly-old",
                metadata,
                header,
                [
                    [
                        "NGDP_R_PA_XDC",
                        "Real Gross Domestic Product",
                        "Tổng sản phẩm theo giá so sánh 2010",
                        "NGDP_R_PA_XDC",
                        "Q",
                        "2010",
                        100,
                        200,
                    ]
                ],
            ),
            (
                "quarterly-current",
                metadata,
                header,
                [
                    [
                        "NGDP_R_PA_XDC",
                        "Real Gross Domestic Product",
                        "Tổng sản phẩm theo giá so sánh 2020",
                        "NGDP_R_PA_XDC",
                        "Q",
                        "2010",
                        110,
                        220,
                    ]
                ],
            ),
        ]
    )

    observations, _releases, warnings = _parse_excel(
        payload,
        dataset=dataset,
        observed_at=datetime(2026, 8, 18, tzinfo=UTC),
        source_url=dataset.excel_url,
        published_at=datetime(2026, 8, 18, tzinfo=UTC),
    )

    assert [item.value for item in observations] == ["110", "220"]
    assert any("conflicting values" in warning for warning in warnings)
    assert any("conflicts with descriptor base year" in warning for warning in warnings)


@pytest.mark.unit
def test_excel_accepts_numeric_zero_multiplier_and_validates_descriptor():
    dataset = NSO_DATASETS[2]
    metadata = [
        ("DATA_DOMAIN", "IND"),
        ("REF_AREA", "VN"),
        ("COUNTERPART_AREA", "_Z"),
        ("UNIT_MULT", 0),
        ("FREQ", "M"),
    ]
    header = ["Country code", "Descriptor", "INDICATOR", "BASE_PER", "2026-07"]
    valid = _workbook_bytes(
        [
            (
                "iip",
                metadata,
                header,
                [["AIP_ISIC4_IX", "Industry (2015=100)", "AIP_ISIC4_IX", "2015", 214.6]],
            )
        ]
    )
    observations, _releases, _warnings = _parse_excel(
        valid,
        dataset=dataset,
        observed_at=datetime(2026, 8, 18, tzinfo=UTC),
        source_url=dataset.excel_url,
        published_at=datetime(2026, 8, 18, tzinfo=UTC),
    )
    assert observations[0].value == "214.6"

    invalid = _workbook_bytes(
        [
            (
                "iip",
                metadata,
                header,
                [["AIP_ISIC4_IX", "Unexpected descriptor", "AIP_ISIC4_IX", "2015", 214.6]],
            )
        ]
    )
    with pytest.raises(ValueError, match="expected descriptor/series missing"):
        _parse_excel(
            invalid,
            dataset=dataset,
            observed_at=datetime(2026, 8, 18, tzinfo=UTC),
            source_url=dataset.excel_url,
            published_at=datetime(2026, 8, 18, tzinfo=UTC),
        )


@pytest.mark.unit
def test_trade_outlier_is_retained_but_generates_quality_warning():
    dataset = NSO_DATASETS[3]
    periods = [(f"2025-{month:02d}", str(10_000 + month)) for month in range(1, 8)]
    periods.append(("2025-08", "100000"))
    payload = _sdmx(
        dataset,
        _series("MET", "TXG_FOB_USD", "M", 6, periods)
        + _series("MET", "TMG_CIF_USD", "M", 6, [(period, "9000") for period, _ in periods])
        + _series(
            "MET",
            "TB_USD",
            "M",
            6,
            [(period, str(float(value) - 9000)) for period, value in periods],
        ),
    )

    observations, _releases, warnings = _parse_sdmx(
        payload,
        dataset=dataset,
        observed_at=datetime.now(UTC),
        source_url=dataset.xml_url,
    )

    assert any(item.value == "100000" for item in observations)
    assert any("exceeds 5x" in warning for warning in warnings)


@pytest.mark.unit
def test_nso_release_discovery_and_parser_use_next_day_publication():
    index = (
        b"""
    <a href="https://www.nso.gov.vn/bai-top/2026/08/bao-cao-thang-bay/">Bao cao</a>
    <span class="archive-issue-date">Ngay dang: 03/08/2026</span>
    <span class="archive-reference-period">Ky tham chieu: 7/2026</span>
    <span class="archive-next-release">Lan cong bo sap toi: 03/09/2026</span>
    """.replace(b"Ngay dang", "Ngày đăng".encode())
        .replace(b"Ky tham chieu", "Kỳ tham chiếu".encode())
        .replace(b"Lan cong bo sap toi", "Lần công bố sắp tới".encode())
    )
    release = _discover_latest_release(index)
    report = """
    <p>Chỉ số sản xuất công nghiệp (IIP) tháng Bảy ước tăng 1,2% so với tháng trước
    và tăng 14,5% so với cùng kỳ năm trước.</p>
    <p>Chỉ số giá tiêu dùng (CPI) tháng Bảy giảm 0,12% so với tháng trước;
    tăng 4,45% so với cùng kỳ năm trước. Lạm phát cơ bản tăng 4,19%.</p>
    <p>Tổng mức bán lẻ hàng hóa và doanh thu dịch vụ tiêu dùng tháng Bảy ước đạt
    669,1 nghìn tỷ đồng, tăng 0,9% so với tháng trước và tăng 14,5% so với cùng kỳ.
    Nếu loại trừ yếu tố giá tăng 7,5%.</p>
    <p>Kim ngạch xuất khẩu hàng hóa tháng Bảy đạt 53,08 tỷ USD.</p>
    <p>Kim ngạch nhập khẩu hàng hóa tháng Bảy đạt 56,67 tỷ USD.</p>
    """.encode()

    observations, releases, warnings = _parse_nso_release(
        report, release=release, observed_at=datetime(2026, 8, 18, tzinfo=UTC)
    )

    mapped = {item.indicator_id: item for item in observations}
    assert mapped["vn_cpi_mom"].value == "-0.12"
    assert mapped["vn_retail_sales_value"].unit_multiplier == 1_000_000_000_000
    assert mapped["vn_exports"].value == "53.08"
    assert mapped["vn_exports"].unit_multiplier == 1_000_000_000
    assert mapped["vn_cpi_yoy"].published_at == datetime(2026, 8, 3, 17, tzinfo=UTC)
    assert releases[0]["reference_period"] == "7/2026"
    assert warnings == []


@pytest.mark.unit
def test_sbv_html_parsers_preserve_missing_fields_as_partial_warnings():
    observed = datetime(2026, 8, 18, 5, tzinfo=UTC)
    fx, _, warnings = _parse_sbv_html(
        "Ngày 17/08/2026 Tỷ giá trung tâm 25.580 VND/USD".encode(),
        source_id="sbv_fx",
        observed_at=observed,
        source_url="https://sbv.gov.vn/vi/t%E1%BB%B7-gi%C3%A1",
    )
    assert warnings == []
    assert fx[0].value == "25580"
    assert fx[0].published_at > datetime(2026, 8, 17, 15, tzinfo=UTC)

    rates, _, warnings = _parse_sbv_html(
        "Ngày 17/08/2026 Lãi suất tái cấp vốn 4,5% Lãi suất tái chiết khấu 3,0%".encode(),
        source_id="sbv_rates",
        observed_at=observed,
        source_url="https://sbv.gov.vn/vi/l%C3%A3i-su%E1%BA%A5t1",
    )
    assert {item.indicator_id for item in rates} == {
        "vn_refinancing_rate",
        "vn_rediscount_rate",
    }
    assert any("interbank_overnight" in warning for warning in warnings)


@pytest.mark.unit
def test_sbv_date_is_bound_to_matched_field_not_unrelated_page_date():
    payload = (
        "Cập nhật chung 31/12/2030 "
        + ("nội dung khác " * 120)
        + "Ngày 17/08/2026 Tỷ giá trung tâm 25.580 VND/USD"
    ).encode()

    observations, _releases, _warnings = _parse_sbv_html(
        payload,
        source_id="sbv_fx",
        observed_at=datetime(2026, 8, 18, 5, tzinfo=UTC),
        source_url="https://sbv.gov.vn/vi/t%E1%BB%B7-gi%C3%A1",
    )

    assert observations[0].period_start.date() == date(2026, 8, 17)


@pytest.mark.unit
def test_gdp_release_uses_explicit_roman_quarter_not_reference_month():
    release = {
        "url": "https://www.nso.gov.vn/bai-top/2026/07/bao-cao-quy-ii/",
        "reference_period": "7/2026",
        "published_at": datetime(2026, 7, 4, tzinfo=UTC),
        "next_release_at": None,
    }
    report = (
        "Tổng sản phẩm trong nước (GDP) quý II năm 2026 ước tính tăng 8,39% so với cùng kỳ."
    ).encode()

    observations, _releases, _warnings = _parse_nso_release(
        report, release=release, observed_at=datetime(2026, 8, 18, tzinfo=UTC)
    )

    gdp = next(item for item in observations if item.indicator_id == "vn_real_gdp_yoy")
    assert gdp.period_start.month == 4
    assert gdp.period_end.month == 6


@pytest.mark.unit
def test_archive_strict_pit_returns_correct_revision_first_seen(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    first = _observation("3.1", datetime(2026, 1, 10, tzinfo=UTC))
    revised = _observation(
        "3.3",
        datetime(2026, 1, 20, tzinfo=UTC),
        published=datetime(2026, 1, 15, tzinfo=UTC),
    )
    archive.record_fetch(_fetch("first", first))
    archive.record_fetch(_fetch("revision", revised))

    early = archive.observations_for_window(
        start=datetime(2025, 1, 1, tzinfo=UTC),
        as_of=datetime(2026, 1, 12, tzinfo=UTC),
        observation_factory=MacroObservation,
    )
    late = archive.observations_for_window(
        start=datetime(2025, 1, 1, tzinfo=UTC),
        as_of=datetime(2026, 1, 25, tzinfo=UTC),
        observation_factory=MacroObservation,
    )

    assert [(item.value, item.first_seen_at) for item in early] == [
        ("3.1", datetime(2026, 1, 10, tzinfo=UTC))
    ]
    assert [(item.value, item.first_seen_at) for item in late] == [
        ("3.3", datetime(2026, 1, 20, tzinfo=UTC))
    ]


@pytest.mark.unit
def test_archive_preserves_a_b_a_revision_occurrences(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    published = datetime(2026, 1, 5, tzinfo=UTC)
    first = _observation("3.1", datetime(2026, 1, 10, tzinfo=UTC), published=published)
    second = _observation("3.3", datetime(2026, 1, 20, tzinfo=UTC), published=published)
    reverted = _observation("3.1", datetime(2026, 1, 30, tzinfo=UTC), published=published)
    for fetch_id, item in (("a1", first), ("b", second), ("a2", reverted)):
        archive.record_fetch(_fetch(fetch_id, item))

    def at(day):
        rows = archive.observations_for_window(
            start=datetime(2025, 1, 1, tzinfo=UTC),
            as_of=datetime(2026, 1, day, tzinfo=UTC),
            observation_factory=MacroObservation,
        )
        return rows[0].value

    assert at(15) == "3.1"
    assert at(25) == "3.3"
    assert at(31) == "3.1"


@pytest.mark.unit
def test_repeat_fetch_does_not_leak_future_retrieved_at(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    original = _observation("3.1", datetime(2026, 1, 10, tzinfo=UTC))
    repeated = replace(
        original,
        first_seen_at=datetime(2026, 2, 10, tzinfo=UTC),
        retrieved_at=datetime(2026, 2, 10, tzinfo=UTC),
    )
    archive.record_fetch(_fetch("original", original))
    archive.record_fetch(_fetch("repeat", repeated))

    rows = archive.observations_for_window(
        start=datetime(2025, 1, 1, tzinfo=UTC),
        as_of=datetime(2026, 1, 20, tzinfo=UTC),
        observation_factory=MacroObservation,
    )

    assert rows[0].retrieved_at == datetime(2026, 1, 10, tzinfo=UTC)


@pytest.mark.unit
def test_source_url_change_does_not_rewrite_historical_provenance(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    original = _observation("3.1", datetime(2026, 1, 10, tzinfo=UTC))
    moved = replace(
        original,
        first_seen_at=datetime(2026, 2, 10, tzinfo=UTC),
        retrieved_at=datetime(2026, 2, 10, tzinfo=UTC),
        source_url="https://nsdp.nso.gov.vn/GSO-chung/SDMXFiles/GSO/CPIVNM-v2.xml",
    )
    archive.record_fetch(_fetch("original-url", original))
    moved_fetch = _fetch("moved-url", moved)
    moved_fetch.source_url = moved.source_url
    archive.record_fetch(moved_fetch)

    rows = archive.observations_for_window(
        start=datetime(2025, 1, 1, tzinfo=UTC),
        as_of=datetime(2026, 1, 20, tzinfo=UTC),
        observation_factory=MacroObservation,
    )

    assert len(rows) == 1
    assert rows[0].source_url == original.source_url


@pytest.mark.unit
def test_archive_excludes_backfill_first_seen_after_as_of(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    observation = _observation(
        "3.1",
        datetime(2026, 8, 18, tzinfo=UTC),
        published=datetime(2025, 12, 31, tzinfo=UTC),
    )
    archive.record_fetch(_fetch("backfill", observation))

    assert (
        archive.observations_for_window(
            start=datetime(2025, 1, 1, tzinfo=UTC),
            as_of=datetime(2026, 1, 10, tzinfo=UTC),
            observation_factory=MacroObservation,
        )
        == []
    )


@pytest.mark.unit
def test_archive_permissions_and_symlink_fail_closed(tmp_path):
    cfg = _config(tmp_path)
    archive = VietnamMacroArchive(cfg.archive_path)
    assert archive.path.stat().st_mode & 0o777 == 0o600

    archive.path.chmod(0o644)
    with pytest.raises(ArchiveConfigurationError, match="safely app-owned"):
        archive.fetch_run_count()


class Response:
    def __init__(self, status_code=200, body=b"ok", headers=None):
        self.status_code = status_code
        self.body = body
        self.headers = headers or {}
        self.closed = False

    def iter_content(self, _size):
        yield self.body

    def close(self):
        self.closed = True


class Session:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = []

    def get(self, url, **kwargs):
        self.calls.append((url, kwargs))
        response = self.responses.pop(0)
        if isinstance(response, BaseException):
            raise response
        return response


@pytest.mark.unit
def test_http_redirect_outside_allowlist_and_oversize_fail_closed(tmp_path):
    redirected = Response(302, headers={"Location": "https://evil.test/data"})
    session = Session([redirected])
    client = VietnamMacroClient(_config(tmp_path), session=session, sleep=lambda _n: None)
    result = client._fetch(
        provider="sbv_html",
        source_id="test",
        source_url="https://sbv.gov.vn/test",
        max_bytes=MAX_HTML_BYTES,
        parser=lambda _body, _observed: ([], [], []),
    )
    assert result.status is MacroStatus.UNAVAILABLE
    assert len(session.calls) == 1
    assert redirected.closed is True

    oversized = Response(body=b"x" * (MAX_HTML_BYTES + 1))
    result = VietnamMacroClient(
        _config(tmp_path), session=Session([oversized]), sleep=lambda _n: None
    )._fetch(
        provider="sbv_html",
        source_id="test",
        source_url="https://sbv.gov.vn/test",
        max_bytes=MAX_HTML_BYTES,
        parser=lambda _body, _observed: ([], [], []),
    )
    assert result.status is MacroStatus.UNAVAILABLE
    assert oversized.closed is True


@pytest.mark.unit
def test_http_retries_three_times_and_honors_retry_after(tmp_path):
    responses = [Response(429, headers={"Retry-After": "1"}) for _ in range(3)]
    responses.append(Response(200, body=b"usable"))
    session = Session(responses)
    sleeps = []
    client = VietnamMacroClient(_config(tmp_path), session=session, sleep=sleeps.append)

    result = client._fetch(
        provider="nso_release",
        source_id="test",
        source_url="https://www.nso.gov.vn/test",
        max_bytes=MAX_HTML_BYTES,
        parser=lambda _body, _observed: ([], [], []),
    )

    assert result.status is MacroStatus.AVAILABLE
    assert len(session.calls) == 4
    assert sleeps == [1.0, 1.0, 1.0]


@pytest.mark.unit
@pytest.mark.parametrize("failure", [Response(500), requests.Timeout("timed out")])
def test_http_5xx_and_timeout_use_initial_plus_three_attempts(tmp_path, failure):
    if isinstance(failure, Response):
        responses = [Response(500) for _ in range(4)]
    else:
        responses = [requests.Timeout("timed out") for _ in range(4)]
    session = Session(responses)
    client = VietnamMacroClient(_config(tmp_path), session=session, sleep=lambda _n: None)

    result = client._fetch(
        provider="nso_release",
        source_id="test",
        source_url="https://www.nso.gov.vn/test",
        max_bytes=MAX_HTML_BYTES,
        parser=lambda _body, _observed: ([], [], []),
    )

    assert result.status is MacroStatus.UNAVAILABLE
    assert len(session.calls) == 4


@pytest.mark.unit
def test_http_304_skips_parser_and_sbv_403_is_explicit(tmp_path):
    parsed = []
    not_modified = VietnamMacroClient(
        _config(tmp_path), session=Session([Response(304)]), sleep=lambda _n: None
    )._fetch(
        provider="nso_sdmx",
        source_id="nso_cpi",
        source_url=NSO_DATASETS[0].xml_url,
        max_bytes=MAX_HTML_BYTES,
        parser=lambda _body, _observed: parsed.append(True),
    )
    assert not_modified.request_succeeded is True
    assert not_modified.http_status == 304
    assert parsed == []

    blocked = VietnamMacroClient(
        _config(tmp_path), session=Session([Response(403)]), sleep=lambda _n: None
    )._fetch(
        provider="sbv_html",
        source_id="sbv_fx",
        source_url="https://sbv.gov.vn/vi/t%E1%BB%B7-gi%C3%A1",
        max_bytes=MAX_HTML_BYTES,
        parser=lambda _body, _observed: ([], [], []),
    )
    assert blocked.http_status == 403
    assert any("WAF/authorization" in warning for warning in blocked.warnings)


@pytest.mark.unit
def test_malformed_xml_fails_closed_without_observations(tmp_path):
    dataset = NSO_DATASETS[0]
    result = VietnamMacroClient(
        _config(tmp_path),
        session=Session([Response(200, body=b"<not-closed")]),
        sleep=lambda _n: None,
    )._fetch(
        provider="nso_sdmx",
        source_id=dataset.source_id,
        source_url=dataset.xml_url,
        max_bytes=MAX_HTML_BYTES,
        parser=lambda body, observed: _parse_sdmx(
            body,
            dataset=dataset,
            observed_at=observed,
            source_url=dataset.xml_url,
        ),
    )

    assert result.status is MacroStatus.UNAVAILABLE
    assert result.observations == []
    assert any("malformed" in warning for warning in result.warnings)


@pytest.mark.unit
def test_service_never_exposes_future_or_late_first_seen_and_returns_partial(tmp_path):
    cfg = _config(tmp_path, providers=("nso_sdmx",))
    archive = VietnamMacroArchive(cfg.archive_path)
    observed = datetime(2026, 1, 10, tzinfo=UTC)
    archive.record_fetch(_fetch("eligible", _observation("3.1", observed)))
    service = VietnamMacroService(cfg, archive=archive)

    unavailable = service.load_evidence(datetime(2026, 1, 9, tzinfo=UTC))
    available = service.load_evidence(datetime(2026, 1, 12, tzinfo=UTC))

    assert unavailable.status is MacroStatus.UNAVAILABLE
    assert unavailable.observations == []
    assert available.status is MacroStatus.PARTIAL
    assert [item.value for item in available.observations] == ["3.1"]
    assert "vn_cpi_yoy" in available.warnings[-1]


@pytest.mark.unit
def test_config_environment_precedence_and_profile_excludes_archive_path(tmp_path, monkeypatch):
    monkeypatch.setenv("TRADINGAGENTS_VN_MACRO_ENABLED", "true")
    monkeypatch.setenv("TRADINGAGENTS_VN_MACRO_PROVIDERS", "nso_sdmx")
    monkeypatch.setenv("TRADINGAGENTS_VN_MACRO_LOOKBACK_MONTHS", "36")
    monkeypatch.setenv(
        "TRADINGAGENTS_VN_MACRO_ARCHIVE_PATH", str(tmp_path / "secret-dir" / "macro.db")
    )
    cfg = VietnamMacroConfig.from_env({"vn_macro": {"enabled": False, "lookback_months": 12}})
    service = VietnamMacroService(cfg)

    assert cfg.enabled is True
    assert cfg.providers == ("nso_sdmx",)
    assert cfg.lookback_months == 36
    assert str(cfg.archive_path) not in service.profile_fingerprint()


@pytest.mark.unit
def test_source_metadata_exposes_singular_fetch_id_for_pipeline(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    observation = _observation("3.1", datetime(2026, 1, 10, tzinfo=UTC))
    archive.record_fetch(_fetch("fetch-one", observation))

    metadata = archive.source_results(("nso_sdmx",), as_of=datetime(2026, 1, 12, tzinfo=UTC))[0]

    assert metadata["fetch_id"] == "fetch-one"
    assert metadata["fetch_ids"] == ["fetch-one"]
    assert metadata["point_in_time_quality"] == "proxy"


@pytest.mark.unit
def test_304_keeps_prior_quality_warning_and_evidence_count(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    observation = _observation("3.1", datetime(2026, 1, 10, tzinfo=UTC))
    initial = _fetch("quality-run", observation)
    initial.status = MacroStatus.PARTIAL
    initial.warnings = ["Quality warning: official source anomaly retained."]
    archive.record_fetch(initial)
    unchanged = _FetchResult(
        fetch_id="not-modified",
        provider="nso_sdmx",
        source_id="nso_cpi",
        source_url=NSO_DATASETS[0].xml_url,
        status=MacroStatus.AVAILABLE,
        started_at=datetime(2026, 1, 11, tzinfo=UTC),
        completed_at=datetime(2026, 1, 11, 0, 0, 1, tzinfo=UTC),
        http_status=304,
        request_succeeded=True,
    )
    archive.record_fetch(unchanged)

    metadata = archive.source_results(("nso_sdmx",), as_of=datetime(2026, 1, 12, tzinfo=UTC))[0]

    assert metadata["fetch_id"] == "not-modified"
    assert metadata["observation_count"] == 1
    assert metadata["status"] == "partial"
    assert "quality-run" in metadata["fetch_ids"]
    assert any("anomaly" in warning for warning in metadata["warnings"])


@pytest.mark.unit
def test_304_baseline_matches_source_url_not_old_excel_fallback(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    seen = datetime(2026, 1, 10, tzinfo=UTC)
    xml_observation = _observation("3.1", seen)
    archive.record_fetch(_fetch("xml-data", xml_observation))

    excel_observation = replace(
        xml_observation,
        value="9.9",
        source_provider="nso_excel",
        source_url=NSO_DATASETS[0].excel_url,
        first_seen_at=seen + timedelta(days=1),
        retrieved_at=seen + timedelta(days=1),
        point_in_time_quality="partial",
    )
    archive.record_fetch(
        _FetchResult(
            fetch_id="excel-fallback",
            provider="nso_sdmx",
            source_id="nso_cpi",
            source_url=NSO_DATASETS[0].excel_url,
            status=MacroStatus.PARTIAL,
            observations=[excel_observation],
            started_at=seen + timedelta(days=1),
            completed_at=seen + timedelta(days=1),
            http_status=200,
            request_succeeded=True,
            warnings=["Excel fallback warning."],
        )
    )
    archive.record_fetch(
        _FetchResult(
            fetch_id="xml-not-modified",
            provider="nso_sdmx",
            source_id="nso_cpi",
            source_url=NSO_DATASETS[0].xml_url,
            status=MacroStatus.AVAILABLE,
            started_at=seen + timedelta(days=2),
            completed_at=seen + timedelta(days=2),
            http_status=304,
            request_succeeded=True,
        )
    )

    metadata = archive.source_results(("nso_sdmx",), as_of=seen + timedelta(days=3))[0]

    assert metadata["fetch_id"] == "xml-not-modified"
    assert metadata["observation_count"] == 1
    assert "xml-data" in metadata["fetch_ids"]
    assert "excel-fallback" not in metadata["fetch_ids"]
    assert all("Excel fallback" not in warning for warning in metadata["warnings"])


@pytest.mark.unit
def test_service_prefers_sdmx_over_newer_excel_for_same_period(tmp_path):
    cfg = _config(tmp_path, providers=("nso_sdmx",))
    archive = VietnamMacroArchive(cfg.archive_path)
    seen = datetime(2026, 1, 10, tzinfo=UTC)
    sdmx = _observation("3.1", seen)
    archive.record_fetch(_fetch("sdmx", sdmx))
    excel = replace(
        sdmx,
        value="9.9",
        published_at=seen + timedelta(days=1),
        first_seen_at=seen + timedelta(days=1),
        retrieved_at=seen + timedelta(days=1),
        source_provider="nso_excel",
        source_url=NSO_DATASETS[0].excel_url,
        point_in_time_quality="partial",
    )
    archive.record_fetch(
        _FetchResult(
            fetch_id="excel",
            provider="nso_sdmx",
            source_id="nso_cpi",
            source_url=NSO_DATASETS[0].excel_url,
            status=MacroStatus.PARTIAL,
            observations=[excel],
            started_at=excel.first_seen_at,
            completed_at=excel.first_seen_at,
            http_status=200,
            request_succeeded=True,
        )
    )

    result = VietnamMacroService(cfg, archive=archive).load_evidence(seen + timedelta(days=2))

    assert result.observations[0].value == "3.1"
    assert result.observations[0].source_provider == "nso_sdmx"


@pytest.mark.unit
def test_growth_deduplicates_excel_and_sdmx_for_the_same_period():
    seen = datetime(2026, 8, 18, tzinfo=UTC)

    def base(period: int, value: str, provider: str) -> MacroObservation:
        return MacroObservation(
            indicator_id="vn_cpi_index",
            value=value,
            unit="index",
            unit_multiplier=1,
            frequency="M",
            period_start=datetime(2026, period, 1, tzinfo=UTC),
            period_end=datetime(
                2026,
                period,
                30 if period == 6 else 31,
                23,
                59,
                tzinfo=UTC,
            ),
            published_at=seen,
            first_seen_at=seen,
            retrieved_at=seen,
            source_provider=provider,
            source_series="PCPI_IX",
            source_url=NSO_DATASETS[0].xml_url,
            provisional=None,
            point_in_time_quality="proxy",
            derived_from=[],
        )

    growth = _derived_growth(
        [
            base(6, "100", "nso_excel"),
            base(6, "101", "nso_sdmx"),
            base(7, "110", "nso_excel"),
            base(7, "111.1", "nso_sdmx"),
        ],
        source_id="vn_cpi_index",
        result_id="vn_cpi_mom",
        lag=1,
    )

    assert growth is not None
    assert growth.value == "10"


@pytest.mark.unit
def test_staleness_uses_nso_release_and_gx_completed_sessions():
    observed = datetime(2026, 8, 18, tzinfo=UTC)
    monthly = _observation("3.1", observed)
    stale_monthly = _apply_staleness(
        monthly,
        datetime(2026, 2, 10, tzinfo=UTC),
        next_release_at=datetime(2026, 2, 5, tzinfo=UTC),
    )
    assert stale_monthly.stale is True

    daily = MacroObservation(
        indicator_id="vn_usd_vnd_central",
        value="25580",
        unit="VND_per_USD",
        unit_multiplier=1,
        frequency="D",
        period_start=datetime(2026, 8, 10, tzinfo=UTC),
        period_end=datetime(2026, 8, 10, 23, 59, tzinfo=UTC),
        published_at=datetime(2026, 8, 11, tzinfo=UTC),
        first_seen_at=observed,
        retrieved_at=observed,
        source_provider="sbv_html",
        source_series="central_exchange_rate",
        source_url="https://sbv.gov.vn/vi/t%E1%BB%B7-gi%C3%A1",
        provisional=None,
        point_in_time_quality="proxy",
        derived_from=[],
    )
    stale_daily = _apply_staleness(
        daily,
        datetime(2026, 8, 14, tzinfo=UTC),
        completed_sessions=[datetime(2026, 8, day, tzinfo=UTC).date() for day in (14, 13, 12, 11)],
    )
    assert stale_daily.stale is True

    unknown_calendar = _apply_staleness(
        daily, datetime(2026, 8, 14, tzinfo=UTC), completed_sessions=None
    )
    assert unknown_calendar.point_in_time_quality == "partial"
    assert any("GX completed-session calendar" in item for item in unknown_calendar.warnings)

    incomplete_calendar = _apply_staleness(
        daily,
        datetime(2026, 8, 14, tzinfo=UTC),
        completed_sessions=[date(2026, 8, 12)],
    )
    assert incomplete_calendar.stale is True
    assert incomplete_calendar.point_in_time_quality == "partial"
    assert any("response was incomplete" in item for item in incomplete_calendar.warnings)


@pytest.mark.unit
def test_release_calendar_correction_is_versioned_by_first_seen(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)

    def release_fetch(fetch_id, seen, next_release):
        return _FetchResult(
            fetch_id=fetch_id,
            provider="nso_release",
            source_id="nso_release_report",
            source_url="https://www.nso.gov.vn/bai-top/2026/01/release/",
            status=MacroStatus.AVAILABLE,
            releases=[
                {
                    "provider": "nso_release",
                    "reference_period": "12/2025",
                    "published_at": datetime(2026, 1, 2, tzinfo=UTC),
                    "next_release_at": next_release,
                    "source_url": "https://www.nso.gov.vn/bai-top/2026/01/release/",
                    "first_seen_at": seen,
                }
            ],
            started_at=seen,
            completed_at=seen,
            http_status=200,
            request_succeeded=True,
        )

    archive.record_fetch(
        release_fetch(
            "release-a",
            datetime(2026, 1, 3, tzinfo=UTC),
            datetime(2026, 2, 3, tzinfo=UTC),
        )
    )
    archive.record_fetch(
        release_fetch(
            "release-b",
            datetime(2026, 1, 10, tzinfo=UTC),
            datetime(2026, 2, 5, tzinfo=UTC),
        )
    )

    early = archive.release_for_period(
        "nso_release", "12/2025", as_of=datetime(2026, 1, 5, tzinfo=UTC)
    )
    late = archive.release_for_period(
        "nso_release", "12/2025", as_of=datetime(2026, 1, 15, tzinfo=UTC)
    )
    assert early["next_release_at"].startswith("2026-02-03")
    assert late["next_release_at"].startswith("2026-02-05")


@pytest.mark.unit
def test_latest_sbv_waf_marks_cached_observation_stale(tmp_path):
    cfg = _config(tmp_path, providers=("sbv_html",))
    archive = VietnamMacroArchive(cfg.archive_path)
    seen = datetime(2026, 8, 11, tzinfo=UTC)
    daily = MacroObservation(
        indicator_id="vn_usd_vnd_central",
        value="25580",
        unit="VND_per_USD",
        unit_multiplier=1,
        frequency="D",
        period_start=datetime(2026, 8, 10, tzinfo=UTC),
        period_end=datetime(2026, 8, 10, 23, 59, tzinfo=UTC),
        published_at=seen,
        first_seen_at=seen,
        retrieved_at=seen,
        source_provider="sbv_html",
        source_series="central_exchange_rate",
        source_url="https://sbv.gov.vn/vi/t%E1%BB%B7-gi%C3%A1",
        provisional=None,
        point_in_time_quality="proxy",
        derived_from=[],
    )
    successful = _FetchResult(
        fetch_id="sbv-ok",
        provider="sbv_html",
        source_id="sbv_fx",
        source_url=daily.source_url,
        status=MacroStatus.AVAILABLE,
        observations=[daily],
        started_at=seen,
        completed_at=seen,
        http_status=200,
        request_succeeded=True,
    )
    failed = _FetchResult(
        fetch_id="sbv-waf",
        provider="sbv_html",
        source_id="sbv_fx",
        source_url=daily.source_url,
        status=MacroStatus.UNAVAILABLE,
        started_at=datetime(2026, 8, 12, tzinfo=UTC),
        completed_at=datetime(2026, 8, 12, tzinfo=UTC),
        http_status=403,
        request_succeeded=False,
        warnings=["Official macro source access was rejected (WAF/authorization)."],
    )
    archive.record_fetch(successful)
    archive.record_fetch(failed)
    service = VietnamMacroService(
        cfg,
        archive=archive,
        completed_session_resolver=lambda _as_of, _count: [date(2026, 8, 11)],
    )

    result = service.load_evidence(datetime(2026, 8, 13, tzinfo=UTC))

    assert result.observations[0].stale is True
    assert any("access-blocked" in warning for warning in result.observations[0].warnings)


@pytest.mark.unit
def test_sbv_endpoint_health_does_not_cross_contaminate_fx_and_rates(tmp_path):
    cfg = _config(tmp_path, providers=("sbv_html",))
    archive = VietnamMacroArchive(cfg.archive_path)
    seen = datetime(2026, 8, 11, tzinfo=UTC)

    def sbv_item(indicator, value, url, series):
        return MacroObservation(
            indicator_id=indicator,
            value=value,
            unit="percent" if "rate" in indicator else "VND_per_USD",
            unit_multiplier=1,
            frequency="D",
            period_start=datetime(2026, 8, 10, tzinfo=UTC),
            period_end=datetime(2026, 8, 10, 23, 59, tzinfo=UTC),
            published_at=seen,
            first_seen_at=seen,
            retrieved_at=seen,
            source_provider="sbv_html",
            source_series=series,
            source_url=url,
            provisional=None,
            point_in_time_quality="proxy",
            derived_from=[],
        )

    fx_url = "https://sbv.gov.vn/vi/t%E1%BB%B7-gi%C3%A1"
    rates_url = "https://sbv.gov.vn/vi/l%C3%A3i-su%E1%BA%A5t1"
    fx = sbv_item("vn_usd_vnd_central", "25580", fx_url, "central_exchange_rate")
    rate = sbv_item("vn_refinancing_rate", "4.5", rates_url, "refinancing_rate")
    for fetch_id, source_id, item in (
        ("fx-ok", "sbv_fx", fx),
        ("rate-ok", "sbv_rates", rate),
    ):
        archive.record_fetch(
            _FetchResult(
                fetch_id=fetch_id,
                provider="sbv_html",
                source_id=source_id,
                source_url=item.source_url,
                status=MacroStatus.AVAILABLE,
                observations=[item],
                started_at=seen,
                completed_at=seen,
                http_status=200,
                request_succeeded=True,
            )
        )
    archive.record_fetch(
        _FetchResult(
            fetch_id="rates-waf",
            provider="sbv_html",
            source_id="sbv_rates",
            source_url=rates_url,
            status=MacroStatus.UNAVAILABLE,
            started_at=datetime(2026, 8, 12, tzinfo=UTC),
            completed_at=datetime(2026, 8, 12, tzinfo=UTC),
            http_status=403,
            request_succeeded=False,
            warnings=["Official macro source access was rejected (WAF/authorization)."],
        )
    )
    service = VietnamMacroService(
        cfg,
        archive=archive,
        completed_session_resolver=lambda _as_of, _count: [date(2026, 8, 11)],
    )

    result = service.load_evidence(datetime(2026, 8, 13, tzinfo=UTC))
    mapped = {item.indicator_id: item for item in result.observations}

    assert mapped["vn_refinancing_rate"].stale is True
    assert mapped["vn_usd_vnd_central"].stale is False


@pytest.mark.unit
def test_live_status_uses_cached_evidence_when_sbv_is_blocked(tmp_path):
    cfg = _config(tmp_path, providers=("sbv_html",))
    archive = VietnamMacroArchive(cfg.archive_path)
    seen = datetime.now(UTC) - timedelta(days=1)
    cached = MacroObservation(
        indicator_id="vn_refinancing_rate",
        value="4.5",
        unit="percent",
        unit_multiplier=1,
        frequency="D",
        period_start=seen.replace(hour=0, minute=0, second=0, microsecond=0),
        period_end=seen.replace(hour=23, minute=59, second=59, microsecond=0),
        published_at=seen,
        first_seen_at=seen,
        retrieved_at=seen,
        source_provider="sbv_html",
        source_series="refinancing_rate",
        source_url="https://sbv.gov.vn/vi/l%C3%A3i-su%E1%BA%A5t1",
        provisional=None,
        point_in_time_quality="proxy",
        derived_from=[],
    )
    archive.record_fetch(
        _FetchResult(
            fetch_id="cached",
            provider="sbv_html",
            source_id="sbv_rates",
            source_url=cached.source_url,
            status=MacroStatus.AVAILABLE,
            observations=[cached],
            started_at=seen,
            completed_at=seen,
            http_status=200,
            request_succeeded=True,
        )
    )
    service = VietnamMacroService(cfg, archive=archive)
    live_failure = _FetchResult(
        fetch_id="live-waf",
        provider="sbv_html",
        source_id="sbv_rates",
        source_url=cached.source_url,
        status=MacroStatus.UNAVAILABLE,
        started_at=datetime.now(UTC),
        completed_at=datetime.now(UTC),
        http_status=403,
        request_succeeded=False,
        warnings=["Official macro source access was rejected (WAF/authorization)."],
    )
    service._collect_provider = lambda _provider, cache=False: [live_failure]

    status = service.status(live=True)

    assert status["status"] == "partial"
    assert status["usable"] is True
    assert status["observation_count"] >= 1
    assert status["issues"] == []


@pytest.mark.unit
def test_render_converts_trade_values_to_billion_usd():
    seen = datetime(2026, 8, 18, tzinfo=UTC)
    trade = MacroObservation(
        indicator_id="vn_exports",
        value="53081.329334",
        unit="USD",
        unit_multiplier=1_000_000,
        frequency="M",
        period_start=datetime(2026, 7, 1, tzinfo=UTC),
        period_end=datetime(2026, 7, 31, tzinfo=UTC),
        published_at=seen,
        first_seen_at=seen,
        retrieved_at=seen,
        source_provider="nso_sdmx",
        source_series="TXG_FOB_USD",
        source_url=NSO_DATASETS[3].xml_url,
        provisional=True,
        point_in_time_quality="proxy",
        derived_from=[],
    )
    rendered = render_vietnam_macro_result(
        VietnamMacroResult(MacroStatus.PARTIAL, seen, [trade], [], [])
    )

    assert "53.0813 | billion USD | 1" in rendered


@pytest.mark.unit
def test_archive_schema_does_not_store_raw_payload_or_secrets(tmp_path):
    archive = VietnamMacroArchive(_config(tmp_path).archive_path)
    with sqlite3.connect(archive.path) as connection:
        schema = " ".join(
            row[0]
            for row in connection.execute("SELECT sql FROM sqlite_master WHERE sql IS NOT NULL")
        )
    assert "raw_payload" not in schema
    assert "api_key" not in schema
